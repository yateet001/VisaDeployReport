import os
from typing import List, Tuple
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from .settings import COMPARISONS_DIR, HIGHLIGHT_COLOR, MATCHING_THRESHOLD, OUTPUT_ROOT, MSTR_OUTPUT_ROOT
from .settings import ensure_dir

def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    return df.astype(str).apply(lambda col: col.str.strip().str.lower())

def compute_row_hash(df: pd.DataFrame) -> pd.Series:
    return pd.util.hash_pandas_object(df, index=False).astype(str)

def get_best_match(row: pd.Series, df_cmp: pd.DataFrame, cols: list):
    row_vals = row[cols].to_numpy(copy=False)
    cmp_vals = df_cmp[cols].to_numpy(copy=False)
    matches = (cmp_vals == row_vals)

    match_counts = matches.sum(axis=1)
    if not match_counts.any():
        return None, [], 0, 0

    best_idx = int(match_counts.argmax())
    best_score = match_counts[best_idx]
    total_cols = len(cols)
    match_pct = (best_score / total_cols) * 100
    matching_cols = np.flatnonzero(matches[best_idx]).tolist()
    return best_idx, matching_cols, best_score, match_pct

def categorize_rows(df1, df2, cols):
    df1_hashes, df2_hashes = set(df1['row_hash']), set(df2['row_hash'])
    df1_status, df1_debug, df2_status, df2_debug = [], [], [], []
    partial_counter, partial_map = 1, {}

    for _, row in df1.iterrows():
        if row['row_hash'] in df2_hashes:
            df1_status.append("Matched"); df1_debug.append(None); continue
        best_idx, _, score, pct = get_best_match(row, df2, cols)
        if score == 0 or (100 - pct) > MATCHING_THRESHOLD:
            df1_status.append("Not Matched"); df1_debug.append(None)
        else:
            debug_id = f"PM_{partial_counter:04d}"
            df1_status.append("Partial Matched"); df1_debug.append(debug_id)
            partial_map[best_idx] = debug_id; partial_counter += 1

    for j, row in df2.iterrows():
        if row['row_hash'] in df1_hashes:
            df2_status.append("Matched"); df2_debug.append(None); continue
        best_idx, _, score, pct = get_best_match(row, df1, cols)
        if score == 0 or (100 - pct) > MATCHING_THRESHOLD:
            df2_status.append("Not Matched"); df2_debug.append(None)
        else:
            debug_id = partial_map.get(j, f"PM_{partial_counter:04d}")
            df2_status.append("Partial Matched"); df2_debug.append(debug_id)
            if j not in partial_map:
                partial_counter += 1

    return df1_status, df2_status, df1_debug, df2_debug

def highlight_cells(ws, df_orig, df_clean, df_cmp, cols):
    highlight = PatternFill(start_color=HIGHLIGHT_COLOR, end_color=HIGHLIGHT_COLOR, fill_type="solid")
    for i, row in enumerate(df_orig.itertuples(index=False), start=2):
        if row.Status not in ["Not Matched", "Partial Matched"]:
            continue
        best_idx, matching_cols, _, _ = get_best_match(df_clean.iloc[i - 2], df_cmp, cols)
        if row.Status == "Partial Matched" and best_idx is not None:
            for col_idx in range(len(cols)):
                if col_idx not in matching_cols:
                    ws.cell(row=i, column=col_idx + 1).fill = highlight
        else:
            for j in range(1, len(cols) + 1):
                ws.cell(row=i, column=j).fill = highlight

def compare_files(file1, file2, output_file):
    """Compare two Excel files and save annotated comparison result.

    - Trims header whitespace
    - Sorts by common columns only (robust to schema differences)
    - Compares only common columns (extras are shown but not used for match)
    """
    print(f"Comparing {file1} vs {file2} -> {output_file}")

    # Read files
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    # Normalize column names (strip whitespace)
    df1.columns = [str(c).strip() for c in df1.columns]
    df2.columns = [str(c).strip() for c in df2.columns]

    # Determine common columns for sort + comparison
    common_cols = [c for c in df1.columns if c in df2.columns]
    if not common_cols:
        # Nothing to compare safely – write both sheets and bail with a clear message
        ensure_dir(os.path.dirname(output_file) or ".")
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            df1.to_excel(writer, sheet_name="MSTR_File", index=False)
            df2.to_excel(writer, sheet_name="PBI_File", index=False)
        print("⚠️ No common columns between files; wrote raw sheets only.")
        return output_file, False

    # Sort for consistency (by common columns only)
    df1 = df1.sort_values(by=common_cols, kind="mergesort").reset_index(drop=True)
    df2 = df2.sort_values(by=common_cols, kind="mergesort").reset_index(drop=True)

    # Clean copies for comparison on common columns
    df1_clean = clean_dataframe(df1[common_cols].copy())
    df2_clean = clean_dataframe(df2[common_cols].copy())

    # Row hashes on common columns
    df1_clean['row_hash'] = compute_row_hash(df1_clean)
    df2_clean['row_hash'] = compute_row_hash(df2_clean)

    # Attach hash to originals (so we can write full frames and still know status)
    df1 = df1.copy()
    df2 = df2.copy()
    df1['row_hash'] = df1_clean['row_hash']
    df2['row_hash'] = df2_clean['row_hash']

    # Status + Debug IDs (based on common cols only)
    df1['Status'], df2['Status'], df1['Debug_ID'], df2['Debug_ID'] = categorize_rows(
        df1_clean, df2_clean, common_cols
    )

    # Reorder columns (put Debug_ID, Status at end)
    def reorder(df):
        base_cols = [c for c in df.columns if c not in ['Debug_ID', 'Status']]
        return df[base_cols + ['Debug_ID', 'Status']]

    df1 = reorder(df1)
    df2 = reorder(df2)

    # Write results
    ensure_dir(os.path.dirname(output_file) or ".")
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df1.to_excel(writer, sheet_name="MSTR_File", index=False)
        df2.to_excel(writer, sheet_name="PBI_File", index=False)

    # Highlight mismatches based on common columns only
    wb = load_workbook(output_file)
    highlight_cells(wb["MSTR_File"], df1, df1_clean, df2_clean, common_cols)
    highlight_cells(wb["PBI_File"], df2, df2_clean, df1_clean, common_cols)

    wb.save(output_file)

    is_fully_matched = df1['Status'].eq('Matched').all() and df2['Status'].eq('Matched').all()
    return output_file, is_fully_matched

def scan_data_tree(root_dir: str):
    """
    <root>/<Report>/<Page>/data/<VisualBase>.xlsx  → mapping[report][page][lower(visualbase)] = abs path
    """
    mapping = {}
    if not os.path.isdir(root_dir):
        return mapping
    for report in os.listdir(root_dir):
        rep_path = os.path.join(root_dir, report)
        if not os.path.isdir(rep_path):
            continue
        mapping.setdefault(report, {})
        for page in os.listdir(rep_path):
            page_path = os.path.join(rep_path, page)
            data_dir = os.path.join(page_path, "data")
            if not os.path.isdir(data_dir):
                continue
            mapping[report].setdefault(page, {})
            for fn in os.listdir(data_dir):
                if fn.lower().endswith((".xlsx", ".csv")):
                    base = os.path.splitext(fn)[0]
                    key = base.lower()
                    mapping[report][page][key] = os.path.join(data_dir, fn)
    return mapping